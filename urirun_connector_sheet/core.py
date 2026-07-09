# Author: Tom Sapletta · https://tom.sapletta.com
# Part of the ifURI solution.
#
# sheet:// connector — first-class URI spreadsheet I/O (XLSX + CSV), so reading a bank
# statement / accounting export and producing an Excel summary no longer needs
# libreoffice or shell. CSV uses the stdlib (always works); XLSX uses openpyxl (pure
# Python, pip — no libreoffice/soffice). Reads return {columns, rows}; writes take rows
# and emit .xlsx or .csv from the extension. Pairs with invoice:///ksef:// — turn the
# parsed invoice/VAT-register rows into an .xlsx an accountant actually opens.

from __future__ import annotations

import csv
import os
from typing import Any

import urirun

from . import _urirun_compat

CONNECTOR_ID = "sheet"
SHEET = _urirun_compat.connector(CONNECTOR_ID, scheme="sheet", target="host", meta={"label": "Spreadsheet I/O (XLSX/CSV)"})


def _read_csv(path: str, max_rows: int) -> dict[str, Any]:
    with open(path, newline="", encoding="utf-8-sig") as fh:
        r = list(csv.reader(fh))
    if not r:
        return {"columns": [], "rows": []}
    cols, body = r[0], r[1:max_rows + 1] if max_rows else r[1:]
    return {"columns": cols, "rows": [dict(zip(cols, row)) for row in body]}


def _read_xlsx(path: str, sheet: str | None, max_rows: int) -> dict[str, Any]:
    from openpyxl import load_workbook  # type: ignore
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = list(next(it))
    except StopIteration:
        return {"columns": [], "rows": [], "sheet": ws.title}
    cols = [str(c) if c is not None else f"col{i}" for i, c in enumerate(header)]
    rows = []
    for n, row in enumerate(it):
        if max_rows and n >= max_rows:
            break
        rows.append({cols[i]: row[i] for i in range(min(len(cols), len(row)))})
    return {"columns": cols, "rows": rows, "sheet": ws.title}


def _write_csv(path: str, columns: list, rows: list) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def _write_xlsx(path: str, columns: list, rows: list, sheet: str, totals: list | None) -> None:
    from openpyxl import Workbook  # type: ignore
    wb = Workbook()
    ws = wb.active
    ws.title = sheet or "Sheet1"
    ws.append(columns)
    for row in rows:
        ws.append([row.get(c) for c in columns])
    if totals:  # a SUM footer over numeric columns named in `totals`
        ws.append([])
        foot = ["TOTAL"] + [""] * (len(columns) - 1)
        for c in totals:
            if c in columns:
                vals = [r.get(c) for r in rows if isinstance(r.get(c), (int, float))]
                foot[columns.index(c)] = round(sum(vals), 2)
        ws.append(foot)
    wb.save(path)


@SHEET.handler("file/query/read", isolated=True,
               meta={"label": "Read an XLSX/CSV into {columns, rows}", "cliAlias": "read"})
def read(path: str = "", sheet: str = "", max_rows: int = 5000) -> dict[str, Any]:
    """Read a spreadsheet at `path` (.xlsx or .csv) into {columns, rows}. `sheet` picks a tab
    for XLSX (default: active). `max_rows` caps the body. CSV needs no extra deps; XLSX needs
    openpyxl (pip install urirun-connector-sheet[xlsx])."""
    path = os.path.expanduser(path)
    if not os.path.isfile(path):
        return {"ok": False, "error": f"file not found: {path}"}
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in (".xlsx", ".xlsm"):
            out = _read_xlsx(path, sheet or None, max_rows)
        elif ext in (".csv", ".tsv", ".txt"):
            out = _read_csv(path, max_rows)
        else:
            return {"ok": False, "error": f"unsupported extension: {ext}"}
    except ImportError:
        return {"ok": False, "error": "XLSX needs openpyxl (pip install openpyxl)", "path": path}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": str(exc), "path": path}
    return {"ok": True, "connector": CONNECTOR_ID, "path": path,
            "columns": out["columns"], "rowCount": len(out["rows"]),
            "sheet": out.get("sheet"), "rows": out["rows"][:1000]}


@SHEET.handler("file/command/write", isolated=True,
               meta={"label": "Write rows to an XLSX/CSV (with optional totals footer)", "cliAlias": "write"})
def write(path: str = "", rows=None, columns=None, sheet: str = "Sheet1", totals=None) -> dict[str, Any]:
    """Write `rows` (list of dicts) to `path`, choosing CSV or XLSX from the extension.
    `columns` sets order (default: keys of the first row); `totals` is a list of numeric
    column names to sum into a footer (XLSX). Creates parent dirs."""
    path = os.path.expanduser(path)
    rows = rows or []
    cols = columns or (list(rows[0].keys()) if rows else [])
    ext = os.path.splitext(path)[1].lower()
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    try:
        if ext in (".xlsx", ".xlsm"):
            _write_xlsx(path, cols, rows, sheet, totals)
        elif ext in (".csv", ".tsv"):
            _write_csv(path, cols, rows)
        else:
            return {"ok": False, "error": f"unsupported extension: {ext} (use .xlsx or .csv)"}
    except ImportError:
        return {"ok": False, "error": "XLSX needs openpyxl (pip install openpyxl)", "path": path}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": str(exc), "path": path}
    return {"ok": True, "connector": CONNECTOR_ID, "path": path, "rows": len(rows),
            "columns": cols, "bytes": os.path.getsize(path) if os.path.exists(path) else 0}


@SHEET.handler("file/query/info", isolated=True, meta={"label": "Sheet names + dimensions of an XLSX"})
def info(path: str = "") -> dict[str, Any]:
    """List the sheets (and their row/col counts) of an XLSX — or columns of a CSV."""
    path = os.path.expanduser(path)
    if not os.path.isfile(path):
        return {"ok": False, "error": f"file not found: {path}"}
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv", ".txt"):
        d = _read_csv(path, 0)
        return {"ok": True, "connector": CONNECTOR_ID, "kind": "csv", "columns": d["columns"], "rowCount": len(d["rows"])}
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(path, read_only=True)
        return {"ok": True, "connector": CONNECTOR_ID, "kind": "xlsx",
                "sheets": [{"name": ws.title, "rows": ws.max_row, "cols": ws.max_column} for ws in wb.worksheets]}
    except ImportError:
        return {"ok": False, "error": "XLSX needs openpyxl (pip install openpyxl)", "path": path}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": str(exc), "path": path}

@SHEET.handler("sheet://host/doctor/query/report", isolated=True, meta={"label": "Connector readiness report"})
def doctor() -> dict[str, Any]:
    """Return a safe, read-only connector readiness report for CI smoke tests."""
    return {
        "ok": True,
        "connector": CONNECTOR_ID,
        "version": _connector_version(),
        "status": "ready",
    }


def _connector_version() -> str:
    try:
        from importlib.metadata import version

        return version("urirun-connector-sheet")
    except Exception:
        return "0.1.0"


def main(argv: list[str] | None = None) -> int:
    return SHEET.cli(argv, manifest_prose=_urirun_compat.load_manifest(__package__))


urirun_bindings = SHEET.bindings
